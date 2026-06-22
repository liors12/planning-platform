"""Excel export of audit findings for architect response workflow.

Reads the same ``audit_results`` dict that :func:`generate_audit_pdf` consumes
and optionally a list of discipline_comments dicts (from the sidecar DB) to
produce a single-sheet RTL Excel workbook. Column schema matches the M8.2
reference file:

  # | חדש? | מקטע בדוח | דיסציפלינה | תא שטח | נושא | סטטוס ממצא |
  תיאור / פעולה נדרשת | סטטוס טיפול | הערות האדריכל

Row sources:
  • בדיקה רב-תחומית      — audit_results["disciplines"]
  • בדיקת תאימות לתב"ע  — audit_results["content"] + m4 sidecar findings
  • הערות מפגישות        — discipline_comments rows (marked ✓ חדש, blue fill)
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ─────────────────────────────────────────────────────────────────────────────
# Brand palette — sourced from compliance_engine/report_generator.py:_CSS
# ─────────────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "FF005030"        # --green-dark (PDF table headers + cover)
COLOR_HEADER_TEXT = "FFFFFFFF"
COLOR_BORDER = "FFD6D6D6"           # --gray-light
COLOR_ALT_ROW = "FFF5F5F5"          # --gray-bg
COLOR_ARCHITECT_BG = "FFFCE4D6"     # peach — architect input columns
COLOR_WARNING_BG = "FFFFF2CC"       # light yellow — read-only banner
COLOR_WARNING_TEXT = "FFC62828"     # --red
COLOR_NEW_ROW_BG = "FFD9EAF7"       # light blue — ✓ חדש meeting-notes rows

FONT_FAMILY = "Arial"


# ─────────────────────────────────────────────────────────────────────────────
# Domain mappings
# ─────────────────────────────────────────────────────────────────────────────
from .constants import DISCIPLINE_NAME_HE, DISCIPLINE_ORDER, STATUS_MAP

DISCIPLINE_HE = DISCIPLINE_NAME_HE
_DISC_SORT_IDX = {DISCIPLINE_NAME_HE.get(k, k): i for i, k in enumerate(DISCIPLINE_ORDER)}

# Sidecar discipline keys → Hebrew labels (mirrors app/sidecar/sidecar/disciplines.py).
_SIDECAR_DISC_HE: dict[str, str] = {
    "sec-3-1":          'שפ"ע — אשפה ופינוי פסולת',
    "sec-3-2":          "גנים ונוף",
    "sec-3-3":          "תשתיות",
    "sec-3-4":          DISCIPLINE_NAME_HE["fire"],
    "sec-3-5":          "ניקוז וחלחול",
    # sec-3-6 (גגות) was merged into arch; comments filed under sec-3-6
    # show the arch label so the Excel section collapses correctly.
    "sec-3-6":          DISCIPLINE_NAME_HE["arch"],
    "sec-3-7":          DISCIPLINE_NAME_HE["arch"],
    "sec-3-8":          "הנחיות סביבתיות",
    "sec-3-9":          "שירותים לדיירים",
    "city-arch":        DISCIPLINE_NAME_HE["city-arch"],
    "public-buildings": DISCIPLINE_NAME_HE["public-buildings"],
    "general":          DISCIPLINE_NAME_HE["general"],
}

STATUS_PRIORITY = {
    "לא תקין":      0,
    "נדרשת השלמה":  1,
    "לא הוגש":      2,
    "הערת פגישה":   2,
    "תקין":         3,
    "לא רלוונטי":  4,
}

SECTION_PRIORITY = {
    "הערות מפגישות":          0,
    "הערות אדריכלית העיר":    1,
    "בדיקה רב-תחומית":        2,
    'בדיקת תאימות לתב"ע':    3,
}

ARCHITECT_STATUS_OPTIONS = ["טופל", "לא טופל", "בטיפול", "לא רלוונטי"]

# Human-readable Hebrew names for well-known sidecar clause_ids that don't
# carry an explicit rule_name_he (e.g. Bug A guard entries from M4 processor).
_SIDECAR_CLAUSE_NAMES: dict[str, str] = {
    "5.table": "טבלת זכויות בנייה",
}


# ─────────────────────────────────────────────────────────────────────────────
# Column layout (matches M8.2 reference schema)
# ─────────────────────────────────────────────────────────────────────────────
COLUMNS: list[tuple[str, int]] = [
    ("#", 6),
    ("חדש?", 8),
    ("מקטע בדוח", 20),
    ("דיסציפלינה", 22),
    ("תא שטח", 10),
    ("נושא", 30),
    ("סטטוס ממצא", 14),
    ("תיאור / פעולה נדרשת", 55),
    ("סטטוס טיפול", 14),
    ("הערות האדריכל", 40),
    ("source_id", 0),   # hidden — stable row identifier for round-trip matching
]
ARCHITECT_COL_INDICES = (9, 10)  # 1-based; col 11 (source_id) is hidden/read-only


# ─────────────────────────────────────────────────────────────────────────────
# Row builders — one per source bucket
# ─────────────────────────────────────────────────────────────────────────────
def _format_plot(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    if s.startswith("plot_"):
        s = s[len("plot_"):]
    return f"תא שטח {s}"


def _join(parts: list[str]) -> str:
    return "\n".join(p for p in parts if p)


def _row_from_discipline(f: dict) -> dict:
    disc_key = f.get("discipline") or ""
    verdict = f.get("verdict") or ""
    return {
        "is_new": False,
        "report_section": "בדיקה רב-תחומית",
        "discipline": DISCIPLINE_HE.get(disc_key, disc_key),
        "plot": "",
        "name": f.get("rule_name_he") or f.get("rule_code") or "",
        "status": STATUS_MAP.get(verdict, verdict),
        "description": _join([f.get("notes_he") or "", f.get("remediation_he") or ""]),
        "source_id": f"disc:{disc_key}:{f.get('rule_code') or ''}",
    }


def _row_from_content(f: dict) -> dict:
    verdict = f.get("verdict") or ""
    ta = f.get("ta_shetach_id") or ""
    return {
        "is_new": False,
        "report_section": 'בדיקת תאימות לתב"ע',
        "discipline": 'בדיקת תאימות לתב"ע',
        "plot": _format_plot(ta),
        "name": f.get("rule_name_he") or f.get("rule_code") or "",
        "status": STATUS_MAP.get(verdict, verdict),
        "description": _join([f.get("notes_he") or "", f.get("remediation_he") or ""]),
        "source_id": f"cont:{f.get('rule_code') or ''}:{ta}",
    }


def _row_from_sidecar(f: dict) -> dict:
    indicator = f.get("compliance_indicator") or ""
    clause_id = f.get("clause_id") or ""
    ta = f.get("ta_shetach_takanon") or ""
    return {
        "is_new": False,
        "report_section": 'בדיקת תאימות לתב"ע',
        "discipline": 'בדיקת תאימות לתב"ע',
        "plot": _format_plot(ta),
        "name": (
            f.get("rule_name_he")
            or _SIDECAR_CLAUSE_NAMES.get(clause_id)
            or (f"סעיף {clause_id}" if clause_id else "")
        ),
        "status": STATUS_MAP.get(indicator, indicator),
        "description": f.get("reasoning") or "",
        "source_id": f"side:{clause_id}:{ta}",
    }


def _row_from_comment(c: dict) -> dict:
    disc_key = c.get("discipline_key") or ""
    disc_label = _SIDECAR_DISC_HE.get(disc_key, disc_key)
    if disc_key == "city-arch":
        section = "הערות אדריכלית העיר"
    else:
        section = "הערות מפגישות"
    return {
        "is_new": True,
        "report_section": section,
        "discipline": disc_label,
        "plot": "",
        "name": c.get("topic_he") or "",
        "status": "הערת פגישה",
        "description": c.get("action_he") or "",
        "source_id": f"cmt:{c.get('id') or ''}",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────
def export_findings_to_excel(
    audit_results: dict,
    output_path: Path,
    report_version: str = "",
    discipline_comments: list[dict] | None = None,
    discipline_filter: list[str] | None = None,
) -> Path:
    """Build the architect-response workbook.

    Layout
    ------
    Row 1: read-only warning banner (merged A1:J1, light yellow, bold red)
    Row 2: column headers (dark green Ness Ziona band, white bold text)
    Row 3+: data rows sorted by section (meeting notes → compliance → plan checks)

    Blue fill marks ✓ חדש rows from meeting notes (discipline_comments).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    disciplines = list(audit_results.get("disciplines") or [])
    content = list(audit_results.get("content") or [])
    sidecar = list(
        (audit_results.get("m4_summary") or {}).get("sidecar_only_findings") or []
    )
    comments_list = list(discipline_comments or [])

    rows: list[dict] = (
        [_row_from_discipline(f) for f in disciplines]
        + [_row_from_content(f) for f in content]
        + [_row_from_sidecar(f) for f in sidecar]
        + [_row_from_comment(c) for c in comments_list]
    )

    if discipline_filter:
        _allowed = set(discipline_filter)
        rows = [r for r in rows if r.get("discipline") in _allowed]

    rows.sort(key=lambda r: (
        SECTION_PRIORITY.get(r["report_section"], 99),
        _DISC_SORT_IDX.get(r["discipline"], 50),
        STATUS_PRIORITY.get(r["status"], 2),
    ))

    # ── Workbook + sheet ────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"הערות סקירה v{report_version}".strip()
    ws.sheet_view.rightToLeft = True

    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        dim = ws.column_dimensions[get_column_letter(col_idx)]
        dim.width = width
        if width == 0:
            dim.hidden = True

    # Reusable styles
    thin = Side(style="thin", color=COLOR_BORDER)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_HEADER_TEXT)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True, readingOrder=2
    )
    data_font = Font(name=FONT_FAMILY, size=10)
    data_align = Alignment(
        horizontal="right", vertical="center", wrap_text=True, readingOrder=2
    )
    alt_fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    new_fill = PatternFill("solid", fgColor=COLOR_NEW_ROW_BG)
    architect_fill = PatternFill("solid", fgColor=COLOR_ARCHITECT_BG)
    warning_fill = PatternFill("solid", fgColor=COLOR_WARNING_BG)
    warning_font = Font(
        name=FONT_FAMILY, size=11, bold=True, color=COLOR_WARNING_TEXT
    )
    warning_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True, readingOrder=2
    )

    WARNING_ROW = 1
    HEADER_ROW = 2
    DATA_START_ROW = 3
    n_cols = len(COLUMNS)

    # ── Row 1: read-only warning (merged, NOT part of filter) ──────────
    today_str = date.today().strftime("%d/%m/%Y")
    warning_text = (
        "אין לשנות או למחוק עמודות ונתונים בגיליון זה. יש למלא אך ורק את העמודות: ״סטטוס טיפול״ ו-״הערות האדריכל״.\n"
        f'השורות המסומנות בכחול ובעמודת ״חדש?״ הן הערות מפגישות שנוספו עד {today_str}. שאר השורות הן ממצאי סקירת התאימות.'
    )
    ws.merge_cells(
        start_row=WARNING_ROW, start_column=1,
        end_row=WARNING_ROW, end_column=n_cols,
    )
    ws.row_dimensions[WARNING_ROW].height = 40
    wc = ws.cell(row=WARNING_ROW, column=1, value=warning_text)
    wc.font = warning_font
    wc.fill = warning_fill
    wc.alignment = warning_align
    for col_idx in range(1, n_cols + 1):
        ws.cell(row=WARNING_ROW, column=col_idx).border = cell_border
        ws.cell(row=WARNING_ROW, column=col_idx).fill = warning_fill

    # ── Row 2: column headers (filter anchor) ───────────────────────────
    ws.row_dimensions[HEADER_ROW].height = 30
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border

    # ── Data rows ───────────────────────────────────────────────────────
    for i, row in enumerate(rows, start=1):
        excel_row = DATA_START_ROW + (i - 1)
        ws.row_dimensions[excel_row].height = 45

        if row["is_new"]:
            row_fill = new_fill
        elif i % 2 == 0:
            row_fill = alt_fill
        else:
            row_fill = None

        values = [
            i,
            "✓ חדש" if row["is_new"] else "",
            row["report_section"],
            row["discipline"],
            row["plot"],
            row["name"],
            row["status"],
            row["description"],
            "",  # architect: status
            "",  # architect: free-text comments
            row.get("source_id", ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=excel_row, column=col_idx, value=v)
            c.font = data_font
            c.alignment = data_align
            c.border = cell_border
            if col_idx in ARCHITECT_COL_INDICES:
                c.fill = architect_fill
            elif row_fill is not None:
                c.fill = row_fill

    last_data_row = DATA_START_ROW + len(rows) - 1

    # ── Summary footer ──────────────────────────────────────────────────
    footer_row = last_data_row + 2
    fc = ws.cell(row=footer_row, column=1, value=f'סה"כ ממצאים: {len(rows)}')
    fc.font = Font(name=FONT_FAMILY, size=11, bold=True)
    fc.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)

    # ── Freeze panes + auto-filter ──────────────────────────────────────
    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = (
        f"A{HEADER_ROW}:{get_column_letter(n_cols)}{last_data_row}"
    )

    # ── Architect status dropdown (column 9) ────────────────────────────
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ARCHITECT_STATUS_OPTIONS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="ערך לא חוקי",
        error="יש לבחור מתוך הרשימה",
    )
    dv.add(f"I{DATA_START_ROW}:I{last_data_row}")
    ws.add_data_validation(dv)

    wb.save(output_path)
    return output_path
