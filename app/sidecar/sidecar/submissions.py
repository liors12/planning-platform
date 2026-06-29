"""Submission endpoints — Phase 2a Module A + engine integration + Phase 2b PDF serving.

  POST   /projects/{project_id}/submissions     — multipart upload (PDF + opt CAD file)
  GET    /projects/{project_id}/submissions     — list submissions for a project
  GET    /submissions/{submission_id}           — full detail
  POST   /submissions/{submission_id}/run-engine — enqueue engine job
  GET    /submissions/{submission_id}/findings  — raw findings JSON
  GET    /submissions/{submission_id}/pdf       — stream PDF with Range support (Phase 2b)
"""
from __future__ import annotations

import json
import logging
import re
import os
import shutil
import subprocess
import sys
import webbrowser
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

log = logging.getLogger(__name__)

from fastapi import APIRouter, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.engine import Engine
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .config import Config
from .engine_bridge import has_schema
from .models import (ArchitectResponse, EmailCorrection, EmailCorrectionRow,
                     Project, ResponseRow, Submission, SubmissionAttachment)
from .queue_worker import EngineQueue
from .storage import StorageError, sanitize_upload_filename, submission_dir


def _audit_outputs_dir(cfg: Config, tava_number: str, version_string: str) -> Path:
    """The canonical per-submission output directory:
    <data_dir>/audit_outputs/<tava>/v<version>/
    Matches _run_render_only(base_dir=cfg.data_dir, output_subdir="audit_outputs").
    Versions in the DB sometimes carry a "v" prefix and sometimes don't —
    we always store on disk with the prefix."""
    ver = version_string if version_string.startswith("v") else f"v{version_string}"
    return cfg.data_dir / "audit_outputs" / tava_number / ver


def _audit_results_path(cfg: Config, tava_number: str, version_string: str) -> Path | None:
    """Return the on-disk audit_results JSON for this submission, preferring
    the sanitized variant (what the approved PDF was rendered from) and
    falling back to the raw M4. None if neither exists."""
    out_dir = _audit_outputs_dir(cfg, tava_number, version_string)
    for leaf in ("audit_results.m4.sanitized.json", "audit_results.m4.json"):
        p = out_dir / leaf
        if p.exists():
            return p
    return None


def _report_pdf_path(cfg: Config, tava_number: str, version_string: str) -> Path:
    """Where the render writes the report PDF. May or may not exist."""
    ver_bare = version_string[1:] if version_string.startswith("v") else version_string
    return _audit_outputs_dir(cfg, tava_number, version_string) / f"audit_report_{ver_bare}.pdf"


def _report_xlsx_path(cfg: Config, tava_number: str, version_string: str) -> Path:
    """Where _run_export_excel writes the architect-response workbook."""
    ver_bare = version_string[1:] if version_string.startswith("v") else version_string
    return _audit_outputs_dir(cfg, tava_number, version_string) / f"הערות_סקירה_v{ver_bare}.xlsx"


def _comparison_xlsx_path(cfg: Config, tava_number: str, version_string: str) -> Path:
    ver_bare = version_string[1:] if version_string.startswith("v") else version_string
    return _audit_outputs_dir(cfg, tava_number, version_string) / f"השוואה_v{ver_bare}.xlsx"


def _comparison_stats_path(cfg: Config, tava_number: str, version_string: str) -> Path:
    ver_bare = version_string[1:] if version_string.startswith("v") else version_string
    return _audit_outputs_dir(cfg, tava_number, version_string) / f"השוואה_v{ver_bare}_stats.json"


# Two routers because the URL grouping crosses prefixes:
_projects_subs_router = APIRouter(prefix="/projects", tags=["submissions"])
_subs_router = APIRouter(prefix="/submissions", tags=["submissions"])


_VALID_WORKFLOW_STAGES = {"draft", "sent", "response_received", "verified"}


class SubmissionOut(BaseModel):
    id: int
    project_id: int
    version_string: str
    status: str
    workflow_stage: str = "draft"
    pdf_path: str
    cad_path: str | None
    findings_json_path: str | None
    uploaded_at: str
    # True iff audit_results.m4.sanitized.json (or .m4.json fallback)
    # exists on disk under cfg.data_dir/audit_outputs/<tava>/v<ver>/.
    # Frontend uses this to show "הפיקי דו״ח" / "הפיקי אקסל" — independent
    # of DB status, so seeded pilots without findings_json_path still
    # surface the buttons.
    has_audit_results: bool = False
    has_report_pdf: bool = False
    has_report_xlsx: bool = False
    # True iff an architect_responses row exists for this submission (B2).
    has_architect_response: bool = False
    # True iff the three-way comparison xlsx has been generated for this revision.
    has_comparison_xlsx: bool = False
    comparison_fixed: int | None = None
    comparison_total_fixable: int | None = None
    # False on win32+frozen, where the full-audit subprocess can't
    # spawn an external Python interpreter (the render-only path is
    # already in-process). Frontend hides/disables "הפעילי את התוכנה"
    # when False to avoid the misleading [WinError 2] SchemaNotFound
    # path. See queue_worker.engine_run_available.
    engine_run_available: bool = True


class JobOut(BaseModel):
    id: str
    job_type: str
    submission_id: int | None
    status: str
    queued_at: str
    started_at: str | None
    completed_at: str | None
    error: str | None


_VERSION_HINT = (
    "version_string must start with an alphanumeric and contain only [A-Za-z0-9._-]. "
    'Examples that work: "v24.3", "24.3", "draft_2026-05-18".'
)


def _stream_upload_to_disk(upload: UploadFile, target: Path) -> None:
    """Stream the upload to disk so we don't load 100+ MB PDFs into memory."""
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("wb") as out:
        # 1 MB chunks — UploadFile already spools past ~1MB by default.
        while True:
            chunk = upload.file.read(1 << 20)
            if not chunk:
                break
            out.write(chunk)


def _sha256_file(path: Path) -> str:
    """Return the hex SHA-256 digest of a file, reading in 1 MB chunks."""
    import hashlib
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1 << 20)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


_OVERLAY_NAMES = ("extracts.json", "discipline_findings.json")


def _archive_overlays(target_dir: Path) -> None:
    """Rename existing engine overlay files in target_dir with a timestamp suffix.

    Called before a new upload lands in an existing directory so the engine's
    previous frozen output is preserved (not deleted) and does not suppress
    re-analysis of the new files.
    """
    from datetime import datetime as _dt, timezone as _tz
    stamp = _dt.now(_tz.utc).strftime("%Y%m%dT%H%M%SZ")
    for name in _OVERLAY_NAMES:
        src = target_dir / name
        if src.exists():
            dst = target_dir / f"{src.stem}.{stamp}{src.suffix}"
            src.rename(dst)
            log.info("archived overlay %s → %s", src.name, dst.name)


class _StageUpdate(BaseModel):
    stage: str


def make_routers(get_engine, cfg: Config, queue: EngineQueue):
    def _session() -> Session:
        return Session(get_engine())

    def _hydrate(sub: Submission, *, has_arch_response: bool = False) -> SubmissionOut:
        """Build a SubmissionOut with the on-disk flags resolved. Project
        tava_number is read off the loaded relationship; safe because each
        route opens its own session and we serialize before exit."""
        tava = sub.project.tava_number
        _cmp_fixed: int | None = None
        _cmp_total: int | None = None
        _stats_path = _comparison_stats_path(cfg, tava, sub.version_string)
        if _stats_path.exists():
            try:
                _s = json.loads(_stats_path.read_text(encoding="utf-8"))
                _cmp_fixed = _s.get("fixed")
                _cmp_total = _s.get("total_fixable")
            except Exception:
                pass
        return SubmissionOut(
            **sub.to_dict(),
            has_audit_results=_audit_results_path(cfg, tava, sub.version_string) is not None,
            has_report_pdf=_report_pdf_path(cfg, tava, sub.version_string).exists(),
            has_report_xlsx=_report_xlsx_path(cfg, tava, sub.version_string).exists(),
            has_architect_response=has_arch_response,
            engine_run_available=True,
            has_comparison_xlsx=_comparison_xlsx_path(cfg, tava, sub.version_string).exists(),
            comparison_fixed=_cmp_fixed,
            comparison_total_fixable=_cmp_total,
        )

    # ── POST /projects/{project_id}/submissions ────────────────────────

    @_projects_subs_router.post(
        "/{project_id}/submissions",
        response_model=SubmissionOut,
        status_code=201,
    )
    async def create_submission(
        project_id: int,
        version_string: str = Form(..., min_length=1, max_length=64),
        pdf: UploadFile = File(...),
        cad_file: UploadFile | None = File(None),
    ) -> SubmissionOut:
        with _session() as sess:
            project = sess.get(Project, project_id)
            if project is None:
                raise HTTPException(404, f"project {project_id} not found")

            # Validate version + prepare target directory upfront so we fail
            # fast before streaming a multi-MB upload. The storage tree is
            # keyed by tava_number (matches the engine's audit_outputs tree).
            try:
                target_dir = submission_dir(cfg, project.tava_number, version_string)
            except StorageError as exc:
                raise HTTPException(422, f"{exc} {_VERSION_HINT}")

            try:
                pdf_leaf = sanitize_upload_filename(pdf.filename or "submission.pdf")
            except StorageError as exc:
                raise HTTPException(422, str(exc))
            pdf_path = target_dir / pdf_leaf

            cad_path: Path | None = None
            if cad_file is not None and cad_file.filename:
                fname = cad_file.filename
                ext = Path(fname).suffix.lower()
                if ext not in (".dwg", ".dxf", ".dwfx"):
                    raise HTTPException(422, "קובץ ה-CAD חייב להיות בפורמט DXF, DWG או DWFX")
                try:
                    cad_leaf = sanitize_upload_filename(fname)
                except StorageError as exc:
                    raise HTTPException(422, str(exc))
                cad_path = target_dir / cad_leaf

            # Archive any existing engine overlays before writing new files,
            # so a re-upload to the same directory doesn't silently reuse
            # frozen results from a prior run.
            _archive_overlays(target_dir)

            # Stream the uploads to disk.
            _stream_upload_to_disk(pdf, pdf_path)
            if cad_file is not None and cad_path is not None:
                _stream_upload_to_disk(cad_file, cad_path)
                # Validate DXF files structurally — catch truncated/corrupt uploads.
                if cad_path.suffix.lower() == ".dxf":
                    try:
                        import ezdxf
                        ezdxf.readfile(str(cad_path))
                    except Exception:
                        cad_path.unlink(missing_ok=True)
                        raise HTTPException(422, "קובץ DXF אינו תקין או פגום")

            # P2-A: write metadata.json so the render path always has it.
            # Originally this file was synthesized by the Phase 2a "Approach
            # B" code (deleted in Phase 2b — see engine_bridge.py history)
            # and never replaced. The render reads two fields from it
            # (submission_version, submission_date); both have empty-string
            # fallbacks, but the file's EXISTENCE used to be a fail-fast
            # gate. Skip writing if a seed/prior version already put one
            # there — never clobber existing metadata.
            metadata_path = target_dir / "metadata.json"
            if not metadata_path.exists():
                bare_version = (version_string[1:] if version_string.startswith("v")
                                else version_string)
                metadata_path.write_text(
                    json.dumps({
                        "plan_number": project.tava_number,
                        "submission_version": bare_version,
                        "submission_date": datetime.now(timezone.utc).date().isoformat(),
                        "file_name": pdf_leaf,
                        "_source": "platform-sidecar",
                    }, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

            pdf_hash = _sha256_file(pdf_path)
            cad_hash = _sha256_file(cad_path) if cad_path else None

            submission = Submission(
                project_id=project_id,
                version_string=version_string,
                status="uploaded",
                pdf_path=str(pdf_path),
                dwg_path=str(cad_path) if cad_path else None,
                pdf_hash=pdf_hash,
                cad_hash=cad_hash,
            )
            sess.add(submission)
            try:
                sess.commit()
            except IntegrityError:
                sess.rollback()
                # P2-B: NO file cleanup. We used to shutil.rmtree(target_dir)
                # here, which wiped seeded metadata.json + any prior successful
                # upload's PDF whenever a user accidentally re-uploaded the
                # same version (Ellen's exact failure on Friday). The bytes
                # we just streamed either overwrite the prior PDF with
                # identical content (same file) or coexist alongside a
                # different one — both cases are recoverable. The DB row
                # already exists; the user can re-upload after deleting the
                # version via the delete-version action (Priority 3).
                raise HTTPException(
                    409,
                    f"submission {version_string!r} already exists for project {project_id}",
                )
            sess.refresh(submission)
            return _hydrate(submission)

    # ── GET /projects/{project_id}/submissions ─────────────────────────

    @_projects_subs_router.get(
        "/{project_id}/submissions",
        response_model=list[SubmissionOut],
    )
    def list_submissions(project_id: int) -> list[SubmissionOut]:
        with _session() as sess:
            project = sess.get(Project, project_id)
            if project is None:
                raise HTTPException(404, f"project {project_id} not found")
            rows = (
                sess.query(Submission)
                .filter(Submission.project_id == project_id)
                .order_by(Submission.uploaded_at.desc(), Submission.id.desc())
                .all()
            )
            sub_ids = {r.id for r in rows}
            arch_ids: set[int] = {
                r[0] for r in sess.query(ArchitectResponse.submission_id)
                .filter(ArchitectResponse.submission_id.in_(sub_ids))
                .all()
            } if sub_ids else set()
            return [_hydrate(r, has_arch_response=r.id in arch_ids) for r in rows]

    # ── GET /submissions/{id} ──────────────────────────────────────────

    @_subs_router.get("/{submission_id}", response_model=SubmissionOut)
    def get_submission(submission_id: int) -> SubmissionOut:
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            has_arch = sess.query(ArchitectResponse).filter_by(
                submission_id=submission_id).first() is not None
            return _hydrate(sub, has_arch_response=has_arch)

    # ── DELETE /submissions/{id} ───────────────────────────────────────
    # Removes the submission row + every file it produced. After delete,
    # the same version_string can be uploaded fresh — Priority 3's whole
    # point: today's manual sqlite3-CLI surgery becomes a button click.
    #
    # FK cascades (db.py:64 sets PRAGMA foreign_keys = ON):
    #   discipline_comments.submission_id  ON DELETE CASCADE
    #   jobs.submission_id                  ON DELETE SET NULL
    # so deleting the parent row cleans dependent comments automatically
    # and leaves job history intact (with NULL submission_id).
    #
    # On-disk: best-effort. File-delete failures (locked files, race
    # with a running render) log a warning but DON'T abort — the DB
    # delete is the source of truth. Re-upload of the same version
    # works because the unique constraint is keyed on (project_id,
    # version_string), not on disk paths.

    @_subs_router.delete("/{submission_id}", status_code=204)
    def delete_submission(submission_id: int):
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string
            try:
                sess.delete(sub)
                sess.commit()
            except Exception as exc:
                sess.rollback()
                log.exception("delete-submission DB step failed for %s", submission_id)
                raise HTTPException(
                    500,
                    detail={
                        "error_type": "DeleteFailed",
                        "error_message": (
                            "Database delete failed. Please close any open "
                            "report/Excel files for this version and try again."
                        ),
                    },
                ) from exc

        # Best-effort file cleanup, logged but never fatal.
        deleted: list[str] = []
        warnings: list[str] = []
        for target in (
            submission_dir(cfg, tava, version),                      # upload folder
            _audit_outputs_dir(cfg, tava, version),                  # derived outputs
        ):
            try:
                if target.exists():
                    shutil.rmtree(target)
                    deleted.append(str(target))
            except OSError as exc:
                warnings.append(f"{target}: {exc}")
                log.warning("delete-submission could not remove %s: %s", target, exc)
        log.info("deleted submission %s (tava=%s ver=%s); folders removed: %s; warnings: %s",
                 submission_id, tava, version, deleted, warnings or "none")
        return Response(status_code=204)

    # NOTE: PDF re-render goes through comments.py's POST /submissions/{id}/render
    # — single source of truth, comments-aware by default. That endpoint
    # also probes audit_results on disk (not the DB findings flag), so it
    # works for seeded pilots without a per-install engine run.

    # ── POST /submissions/{id}/export-excel ────────────────────────────

    @_subs_router.post(
        "/{submission_id}/export-excel",
        response_model=JobOut,
        status_code=202,
    )
    def export_excel(
        submission_id: int,
        discipline: list[str] = Query(default=[]),
    ) -> JobOut:
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            if _audit_results_path(cfg, tava, sub.version_string) is None:
                raise HTTPException(
                    409,
                    f"submission {submission_id} has no audit_results.m4.json on disk; "
                    "run the engine first.",
                )
        discipline_filter = discipline if discipline else None
        job = queue.enqueue_excel(submission_id, discipline_filter=discipline_filter)
        return JobOut(**job.to_dict())

    # ── GET /submissions/{id}/report.pdf ───────────────────────────────
    # Streams the engine's generated report PDF (distinct from the
    # /pdf endpoint above, which serves the original submission upload).

    @_subs_router.get("/{submission_id}/report.pdf")
    def get_report_pdf(submission_id: int):
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string
        path = _report_pdf_path(cfg, tava, version)
        if not path.exists():
            raise HTTPException(404, f"report PDF not generated yet for submission {submission_id}")
        return Response(
            content=path.read_bytes(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'inline; filename="{path.name}"',
                "Cache-Control": "private, max-age=0, must-revalidate",
            },
        )

    # ── GET /submissions/{id}/report.xlsx ──────────────────────────────

    @_subs_router.get("/{submission_id}/report.xlsx")
    def get_report_xlsx(submission_id: int):
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string
        path = _report_xlsx_path(cfg, tava, version)
        if not path.exists():
            raise HTTPException(404, f"Excel report not generated yet for submission {submission_id}")
        # Hebrew filename: RFC 5987 encoded for cross-browser correctness.
        from urllib.parse import quote
        ascii_name = f"audit_report_v{version.lstrip('v')}.xlsx"
        utf8_name = quote(path.name)
        return Response(
            content=path.read_bytes(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{ascii_name}"; '
                    f"filename*=UTF-8''{utf8_name}"
                ),
                "Cache-Control": "private, max-age=0, must-revalidate",
            },
        )

    # ── POST /submissions/{id}/open-output + /reveal-output ────────────
    # Today Ellen sees "nothing" when she clicks "פתחי דו״ח" because the
    # PDF link is target=_blank — and the Tauri webview swallows _blank
    # navigations. These endpoints solve that without adding a Tauri
    # plugin: the sidecar already has OS access via Python, so it just
    # spawns Explorer / the default app pointed at the generated file.
    #
    # Security: caller picks `kind ∈ {pdf,xlsx}` only. The path is built
    # server-side from cfg.data_dir + tava + version + the canonical
    # leaf name. Never accepts a user-supplied path.

    def _resolve_output(submission_id: int, kind: str) -> Path:
        if kind not in ("pdf", "xlsx"):
            raise HTTPException(422, f"unknown output kind: {kind!r}")
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string
        path = (_report_pdf_path(cfg, tava, version) if kind == "pdf"
                else _report_xlsx_path(cfg, tava, version))
        if not path.exists():
            raise HTTPException(
                404,
                f"{kind} report not generated yet for submission {submission_id}",
            )
        return path

    def _spawn(args: list[str]) -> None:
        """Fire-and-forget. Errors swallowed because the OS file-open
        call shouldn't block or fail the HTTP response — the worst case
        is "user clicks button, nothing happens", which the front-end
        can re-prompt for. Errors do flow to errors.log via the warning."""
        try:
            subprocess.Popen(args, stdin=subprocess.DEVNULL,
                             stdout=subprocess.DEVNULL,
                             stderr=subprocess.DEVNULL)
        except Exception as exc:
            log.warning("OS open failed for %s: %s", args, exc)

    @_subs_router.post("/{submission_id}/open-output", status_code=204)
    def open_output(submission_id: int, kind: str):
        """Open the generated PDF or XLSX in the OS default app."""
        path = _resolve_output(submission_id, kind)
        if sys.platform == "win32":
            os.startfile(str(path))      # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            _spawn(["open", str(path)])
        else:
            _spawn(["xdg-open", str(path)])
        return Response(status_code=204)

    @_subs_router.post("/{submission_id}/reveal-output", status_code=204)
    def reveal_output(submission_id: int, kind: str):
        """Open the file's containing folder in the OS file manager,
        with the file selected/highlighted where the OS supports it."""
        path = _resolve_output(submission_id, kind)
        if sys.platform == "win32":
            # explorer /select highlights the file inside its folder.
            _spawn(["explorer", "/select,", str(path)])
        elif sys.platform == "darwin":
            _spawn(["open", "-R", str(path)])
        else:
            _spawn(["xdg-open", str(path.parent)])
        return Response(status_code=204)

    @_subs_router.post("/open-url", status_code=204)
    def open_url(url: str = Query(...)):
        """Open an external URL in the OS default browser.
        Needed because the Tauri webview ignores target="_blank"."""
        webbrowser.open(url)
        return Response(status_code=204)

    # ── PATCH /submissions/{id}/stage ─────────────────────────────────

    @_subs_router.patch("/{submission_id}/stage", response_model=SubmissionOut)
    def set_workflow_stage(submission_id: int, body: _StageUpdate) -> SubmissionOut:
        """Advance or roll back the workflow stage for a submission."""
        if body.stage not in _VALID_WORKFLOW_STAGES:
            raise HTTPException(
                422, f"invalid stage '{body.stage}'. "
                f"Valid values: {sorted(_VALID_WORKFLOW_STAGES)}"
            )
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            sub.workflow_stage = body.stage
            sess.commit()
            sess.refresh(sub)
            return _hydrate(sub)

    # ── POST /submissions/{id}/upload-response ─────────────────────────
    # Accepts the architect's filled-in Excel, parses source_id/treatment_
    # status/notes from every data row, and upserts into architect_responses
    # + response_rows. Re-uploading replaces the prior response entirely
    # (cascade). Stage auto-flips to response_received on success.

    @_subs_router.post("/{submission_id}/upload-response", response_model=SubmissionOut)
    async def upload_architect_response(
        submission_id: int,
        xlsx: UploadFile = File(...),
    ) -> SubmissionOut:
        fname = xlsx.filename or "response.xlsx"
        if not fname.lower().endswith(".xlsx"):
            raise HTTPException(422, "יש להעלות קובץ Excel בפורמט .xlsx בלבד")

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string

        ver_bare = version[1:] if version.startswith("v") else version
        out_dir = _audit_outputs_dir(cfg, tava, version)
        out_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = out_dir / f"תשובת_אדריכל_v{ver_bare}.xlsx"
        _stream_upload_to_disk(xlsx, xlsx_path)

        try:
            parsed_rows = _parse_response_xlsx(xlsx_path)
        except ValueError as exc:
            raise HTTPException(422, str(exc))
        except Exception as exc:
            log.exception("response xlsx parse failed for submission %s", submission_id)
            raise HTTPException(422, "קובץ האקסל אינו תקין או אינו בפורמט הצפוי") from exc

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            existing = sess.query(ArchitectResponse).filter_by(
                submission_id=submission_id).first()
            if existing is not None:
                sess.delete(existing)
                sess.flush()
            arch_resp = ArchitectResponse(
                submission_id=submission_id,
                xlsx_path=str(xlsx_path),
                row_count=len(parsed_rows),
            )
            sess.add(arch_resp)
            sess.flush()
            for row in parsed_rows:
                sess.add(ResponseRow(
                    response_id=arch_resp.id,
                    source_id=row["source_id"],
                    topic_he=row.get("topic_he") or None,
                    finding_status=row.get("finding_status") or None,
                    description=row.get("description") or None,
                    treatment_status=row.get("treatment_status") or None,
                    architect_notes=row.get("architect_notes") or None,
                ))
            sub.workflow_stage = "response_received"
            sess.commit()
            sess.refresh(sub)
            result = _hydrate(sub, has_arch_response=True)

        log.info("response uploaded for submission %s: %d rows parsed", submission_id,
                 len(parsed_rows))
        return result

    # ── GET /submissions/{id}/response ─────────────────────────────────

    class _ResponseInfo(BaseModel):
        submission_id: int
        row_count: int
        uploaded_at: str
        rows: list[dict]

    @_subs_router.get("/{submission_id}/response", response_model=_ResponseInfo)
    def get_architect_response(submission_id: int) -> _ResponseInfo:
        with _session() as sess:
            arch = sess.query(ArchitectResponse).filter_by(
                submission_id=submission_id).first()
            if arch is None:
                raise HTTPException(404,
                    f"no architect response uploaded for submission {submission_id}")
            rows = [
                {
                    "source_id": r.source_id,
                    "topic_he": r.topic_he,
                    "finding_status": r.finding_status,
                    "description": r.description,
                    "treatment_status": r.treatment_status,
                    "architect_notes": r.architect_notes,
                }
                for r in arch.rows
            ]
            return _ResponseInfo(
                submission_id=submission_id,
                row_count=arch.row_count,
                uploaded_at=arch.uploaded_at.isoformat(),
                rows=rows,
            )

    # ── POST /submissions/{id}/attachments ─────────────────────────────
    # Upload an arbitrary file attachment to a submission. Files are stored
    # under <data_dir>/projects/{tava}/submissions/v{bare}/attachments/.

    class _AttachmentOut(BaseModel):
        id: int
        submission_id: int
        filename: str
        file_size: int
        uploaded_at: str

    @_subs_router.post(
        "/{submission_id}/attachments",
        response_model=_AttachmentOut,
        status_code=201,
    )
    async def upload_attachment(
        submission_id: int,
        file: UploadFile = File(...),
    ) -> _AttachmentOut:
        fname = file.filename or "attachment"
        try:
            safe_name = sanitize_upload_filename(fname)
        except StorageError as exc:
            raise HTTPException(422, str(exc)) from exc

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string

        att_dir = submission_dir(cfg, tava, version) / "attachments"
        att_dir.mkdir(parents=True, exist_ok=True)
        dest = att_dir / safe_name
        # Avoid silent overwrites: suffix the name if a file already exists.
        counter = 1
        stem = dest.stem
        suffix = dest.suffix
        while dest.exists():
            dest = att_dir / f"{stem}_{counter}{suffix}"
            counter += 1

        _stream_upload_to_disk(file, dest)
        file_size = dest.stat().st_size

        with _session() as sess:
            att = SubmissionAttachment(
                submission_id=submission_id,
                filename=dest.name,
                file_path=str(dest),
                file_size=file_size,
            )
            sess.add(att)
            sess.commit()
            sess.refresh(att)
            return _AttachmentOut(
                id=att.id,
                submission_id=att.submission_id,
                filename=att.filename,
                file_size=att.file_size,
                uploaded_at=att.uploaded_at.isoformat(),
            )

    # ── GET /submissions/{id}/attachments ──────────────────────────────

    @_subs_router.get("/{submission_id}/attachments", response_model=list[_AttachmentOut])
    def list_attachments(submission_id: int) -> list[_AttachmentOut]:
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            atts = (
                sess.query(SubmissionAttachment)
                .filter_by(submission_id=submission_id)
                .order_by(SubmissionAttachment.uploaded_at.asc())
                .all()
            )
            return [
                _AttachmentOut(
                    id=a.id,
                    submission_id=a.submission_id,
                    filename=a.filename,
                    file_size=a.file_size,
                    uploaded_at=a.uploaded_at.isoformat(),
                )
                for a in atts
            ]

    # ── DELETE /submissions/{id}/attachments/{att_id} ──────────────────

    @_subs_router.delete("/{submission_id}/attachments/{att_id}", status_code=204)
    def delete_attachment(submission_id: int, att_id: int) -> None:
        with _session() as sess:
            att = sess.get(SubmissionAttachment, att_id)
            if att is None or att.submission_id != submission_id:
                raise HTTPException(404, f"attachment {att_id} not found")
            file_path = Path(att.file_path)
            sess.delete(att)
            sess.commit()
        if file_path.exists():
            try:
                file_path.unlink()
            except OSError:
                log.warning("could not delete attachment file %s", file_path)

    # ── POST /submissions/{id}/run-engine ──────────────────────────────

    @_subs_router.post("/{submission_id}/run-engine", response_model=JobOut, status_code=202)
    def run_engine(submission_id: int) -> JobOut:
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            project = sub.project
            if not has_schema(project.tava_number):
                raise HTTPException(
                    409,
                    f"project {project.id} (tava {project.tava_number}) has no schema file; "
                    "engine cannot run. Phase 3 will add a schema-upload UI.",
                )
            if not Path(sub.pdf_path).exists():
                raise HTTPException(
                    409,
                    f"submission {submission_id} PDF missing from disk at {sub.pdf_path}",
                )
            if sub.pdf_hash:
                duplicate = (
                    sess.query(Submission)
                    .filter(
                        Submission.project_id == sub.project_id,
                        Submission.id != submission_id,
                        Submission.pdf_hash == sub.pdf_hash,
                    )
                    .first()
                )
                if duplicate:
                    raise HTTPException(
                        409,
                        "הקובץ זהה לגרסה הקודמת. לא בוצע ניתוח חוזר.",
                    )

        # Enqueue from outside the session so the new Session inside enqueue_run_audit
        # doesn't conflict.
        job = queue.enqueue_run_audit(submission_id)
        return JobOut(**job.to_dict())

    # ── GET /submissions/{id}/suggest-revision-version ─────────────────

    @_subs_router.get("/{submission_id}/suggest-revision-version")
    def suggest_revision_version(submission_id: int) -> dict:
        """Return a suggested version string for a revision of this submission.
        Increments the last numeric segment: 'v24.3' → 'v24.4'."""
        import re
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            current = sub.version_string
        m = re.match(r"^(.*?)(\d+)$", current)
        if m:
            suggested = f"{m.group(1)}{int(m.group(2)) + 1}"
        else:
            suggested = f"{current}-r2"
        return {"suggested": suggested}

    # ── POST /submissions/{id}/create-revision ─────────────────────────

    @_subs_router.post("/{submission_id}/create-revision",
                       response_model=SubmissionOut, status_code=201)
    async def create_revision(
        submission_id: int,
        version_string: str = Form(..., min_length=1, max_length=64),
        pdf: UploadFile | None = File(None),
        cad_file: UploadFile | None = File(None),
    ) -> SubmissionOut:
        """Create a new submission as a revision of an existing one.
        If no PDF is supplied, copies the source submission's PDF. If no CAD is
        supplied and the source had one, copies it too. Archives any existing
        engine overlays in the new version's directory."""
        with _session() as sess:
            src = sess.get(Submission, submission_id)
            if src is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            project = src.project
            src_pdf_path = Path(src.pdf_path)
            src_cad_path = Path(src.dwg_path) if src.dwg_path else None

            try:
                target_dir = submission_dir(cfg, project.tava_number, version_string)
            except StorageError as exc:
                raise HTTPException(422, f"{exc} {_VERSION_HINT}")

            # Prevent overwriting an existing version — DB will also catch it,
            # but fail fast here with a clearer message.
            existing = (
                sess.query(Submission)
                .filter(Submission.project_id == project.id,
                        Submission.version_string == version_string)
                .first()
            )
            if existing:
                raise HTTPException(
                    409,
                    f"גרסה {version_string!r} כבר קיימת. בחרי מספר גרסה אחר."
                )

        _archive_overlays(target_dir)

        # Determine the new PDF path.
        if pdf is not None and pdf.filename:
            try:
                pdf_leaf = sanitize_upload_filename(pdf.filename)
            except StorageError as exc:
                raise HTTPException(422, str(exc))
            new_pdf_path = target_dir / pdf_leaf
            _stream_upload_to_disk(pdf, new_pdf_path)
        else:
            # Copy source PDF into the new version directory.
            if not src_pdf_path.exists():
                raise HTTPException(
                    409,
                    "קובץ ה-PDF של הגרסה המקורית חסר מהדיסק. יש להעלות קובץ חדש."
                )
            new_pdf_path = target_dir / src_pdf_path.name
            target_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src_pdf_path, new_pdf_path)

        # Determine the new CAD path.
        new_cad_path: Path | None = None
        if cad_file is not None and cad_file.filename:
            fname = cad_file.filename
            ext = Path(fname).suffix.lower()
            if ext not in (".dwg", ".dxf", ".dwfx"):
                raise HTTPException(422, "קובץ ה-CAD חייב להיות בפורמט DXF, DWG או DWFX")
            try:
                cad_leaf = sanitize_upload_filename(fname)
            except StorageError as exc:
                raise HTTPException(422, str(exc))
            new_cad_path = target_dir / cad_leaf
            _stream_upload_to_disk(cad_file, new_cad_path)
            if new_cad_path.suffix.lower() == ".dxf":
                try:
                    import ezdxf
                    ezdxf.readfile(str(new_cad_path))
                except Exception:
                    new_cad_path.unlink(missing_ok=True)
                    raise HTTPException(422, "קובץ DXF אינו תקין או פגום")
        elif src_cad_path and src_cad_path.exists():
            new_cad_path = target_dir / src_cad_path.name
            shutil.copy2(src_cad_path, new_cad_path)

        # Write metadata.json for the new version.
        metadata_path = target_dir / "metadata.json"
        if not metadata_path.exists():
            bare_version = (version_string[1:] if version_string.startswith("v")
                            else version_string)
            metadata_path.write_text(
                json.dumps({
                    "plan_number": project.tava_number,
                    "submission_version": bare_version,
                    "submission_date": datetime.now(timezone.utc).date().isoformat(),
                    "file_name": new_pdf_path.name,
                    "_source": "platform-sidecar-revision",
                }, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        pdf_hash = _sha256_file(new_pdf_path)
        cad_hash = _sha256_file(new_cad_path) if new_cad_path else None

        with _session() as sess:
            revision = Submission(
                project_id=project.id,
                version_string=version_string,
                status="uploaded",
                pdf_path=str(new_pdf_path),
                dwg_path=str(new_cad_path) if new_cad_path else None,
                pdf_hash=pdf_hash,
                cad_hash=cad_hash,
                source_submission_id=submission_id,
            )
            sess.add(revision)
            try:
                sess.commit()
            except IntegrityError:
                sess.rollback()
                raise HTTPException(
                    409,
                    f"גרסה {version_string!r} כבר קיימת. בחרי מספר גרסה אחר."
                )
            sess.refresh(revision)
            return _hydrate(revision)

    # ── POST /submissions/{id}/comparison-excel ────────────────────────

    class _JobOut(BaseModel):
        id: str
        job_type: str
        submission_id: int | None
        status: str
        queued_at: str
        started_at: str | None
        completed_at: str | None
        error: str | None

    @_subs_router.post("/{submission_id}/comparison-excel", response_model=_JobOut, status_code=202)
    def enqueue_comparison_excel(submission_id: int) -> _JobOut:
        """Enqueue a three-way comparison Excel job for a revision submission."""
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            if sub.source_submission_id is None:
                raise HTTPException(
                    422,
                    "הגשה זו אינה גרסה מתוקנת ואינה ניתנת להשוואה."
                )
        try:
            job = queue.enqueue_comparison_excel(submission_id)
        except ValueError as exc:
            raise HTTPException(422, str(exc)) from exc
        return _JobOut(
            id=job.id,
            job_type=job.job_type,
            submission_id=job.submission_id,
            status=job.status,
            queued_at=job.queued_at.isoformat() if job.queued_at else "",
            started_at=job.started_at.isoformat() if job.started_at else None,
            completed_at=job.completed_at.isoformat() if job.completed_at else None,
            error=job.error_json,
        )

    @_subs_router.post("/{submission_id}/open-comparison", status_code=204)
    def open_comparison(submission_id: int) -> Response:
        """Open the comparison xlsx in the OS default app."""
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            tava = sub.project.tava_number
            version = sub.version_string
        path = _comparison_xlsx_path(cfg, tava, version)
        if not path.exists():
            raise HTTPException(404, "קובץ ההשוואה עדיין לא הופק")
        if sys.platform == "win32":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            _spawn(["open", str(path)])
        else:
            _spawn(["xdg-open", str(path)])
        return Response(status_code=204)

    # ── GET /submissions/{id}/findings ─────────────────────────────────

    @_subs_router.get("/{submission_id}/findings")
    def get_findings(submission_id: int):
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            if sub.findings_json_path is None:
                raise HTTPException(
                    409,
                    f"submission {submission_id} has no findings yet (status={sub.status!r}). "
                    "POST /submissions/{id}/run-engine first.",
                )
            findings_file = Path(sub.findings_json_path)
            if not findings_file.exists():
                raise HTTPException(
                    500,
                    f"findings_json_path is recorded but file missing on disk: {findings_file}",
                )
            return json.loads(findings_file.read_text(encoding="utf-8"))

    # ── GET /submissions/{id}/pdf  (Phase 2b — Range-aware streaming) ──
    #
    # react-pdf reads the file with byte-range requests, fetching pdf.js
    # metadata + xref tables before pulling page bodies. Without 206 Partial
    # Content support, the browser falls back to downloading the whole file
    # (100 MB for v24.3) before showing page 1.
    #
    # Implementation notes:
    #   - Starlette's FileResponse offers Range support in recent versions,
    #     but behavior across starlette/uvicorn version combos has been
    #     uneven. We do it explicitly to be deterministic.
    #   - Spec § 8: 127.0.0.1-only, no auth. Same constraint as the rest of
    #     the API.
    #   - Memory: yield in 256 KB chunks. A 100 MB file at 200 chunks max is
    #     trivial; never holds the whole PDF in RAM.

    _PDF_CHUNK_SIZE = 256 * 1024  # 256 KB

    @_subs_router.head("/{submission_id}/pdf")
    def head_pdf(submission_id: int):
        """HEAD: probe for Accept-Ranges + Content-Length without body.
        pdf.js (via react-pdf) sometimes preflights with HEAD before issuing
        the first ranged GET, so we expose it explicitly."""
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            pdf_path = Path(sub.pdf_path)
        if not pdf_path.exists():
            # Soft 404: the row exists but the bytes don't (e.g. user
            # cleared cfg.data_dir, or storage was migrated and this row
            # still points at the old layout). UI can react with a
            # "re-upload required" prompt instead of erroring loudly.
            raise HTTPException(
                404,
                f"submission {submission_id} PDF missing from disk "
                f"(stale file_path: {sub.pdf_path}). Re-upload required.",
            )
        file_size = pdf_path.stat().st_size
        return Response(
            status_code=200,
            headers={
                "Accept-Ranges": "bytes",
                "Content-Type": "application/pdf",
                "Content-Length": str(file_size),
                "Cache-Control": "private, max-age=0, must-revalidate",
            },
        )

    @_subs_router.get("/{submission_id}/pdf")
    def get_pdf(
        submission_id: int,
        request: Request,
        range_header: str | None = Header(None, alias="range"),
    ):
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            pdf_path = Path(sub.pdf_path)
        if not pdf_path.exists():
            # Soft 404: the row exists but the bytes don't (e.g. user
            # cleared cfg.data_dir, or storage was migrated and this row
            # still points at the old layout). UI can react with a
            # "re-upload required" prompt instead of erroring loudly.
            raise HTTPException(
                404,
                f"submission {submission_id} PDF missing from disk "
                f"(stale file_path: {sub.pdf_path}). Re-upload required.",
            )

        file_size = pdf_path.stat().st_size
        filename = pdf_path.name
        common_headers = {
            "Accept-Ranges": "bytes",
            "Content-Type": "application/pdf",
            "Cache-Control": "private, max-age=0, must-revalidate",
            "Content-Disposition": f'inline; filename="{filename}"',
        }

        if range_header is None:
            return StreamingResponse(
                _stream_file_range(pdf_path, 0, file_size - 1, _PDF_CHUNK_SIZE),
                status_code=200,
                headers={**common_headers, "Content-Length": str(file_size)},
            )

        # Parse "Range: bytes=START-END" (single range only — pdf.js never
        # requests multipart ranges).
        rng = _parse_range(range_header, file_size)
        if rng is None:
            return Response(
                status_code=416,  # Requested Range Not Satisfiable
                headers={**common_headers, "Content-Range": f"bytes */{file_size}"},
            )
        start, end = rng
        length = end - start + 1
        return StreamingResponse(
            _stream_file_range(pdf_path, start, end, _PDF_CHUNK_SIZE),
            status_code=206,  # Partial Content
            headers={
                **common_headers,
                "Content-Length": str(length),
                "Content-Range": f"bytes {start}-{end}/{file_size}",
            },
        )

    # ── POST /submissions/{id}/upload-email-response ──────────────────
    # Upload an architect's email-as-PDF. Extracts structured corrections
    # (page, description, category) via Claude and stores them in DB.
    # Re-uploading replaces the previous EmailCorrection for this submission.

    _MAX_EMAIL_PDF_BYTES = 20 * 1024 * 1024  # 20 MB

    class _EmailCorrectionRowOut(BaseModel):
        id: int
        correction_id: int
        page_number: int | None
        change_he: str
        category: str

    class _EmailCorrectionOut(BaseModel):
        id: int
        submission_id: int
        row_count: int
        used_ai: bool
        uploaded_at: str
        rows: list[_EmailCorrectionRowOut]
        raw_text: str | None = None
        error: str | None = None
        error_message: str | None = None

    @_subs_router.post(
        "/{submission_id}/upload-email-response",
        response_model=_EmailCorrectionOut,
        status_code=201,
    )
    async def upload_email_response(
        submission_id: int,
        pdf: UploadFile = File(...),
    ) -> _EmailCorrectionOut:
        """Extract architect corrections from an email-as-PDF and store them."""
        from .email_extract import extract_email_corrections  # noqa: PLC0415

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")

        content = await pdf.read(_MAX_EMAIL_PDF_BYTES + 1)
        if len(content) > _MAX_EMAIL_PDF_BYTES:
            raise HTTPException(
                413,
                f"קובץ ה-PDF גדול מדי — מקסימום {_MAX_EMAIL_PDF_BYTES // (1024*1024)} MB",
            )

        result = extract_email_corrections(content)

        with _session() as sess:
            # Replace any existing EmailCorrection for this submission.
            existing = sess.query(EmailCorrection).filter_by(
                submission_id=submission_id
            ).first()
            if existing:
                sess.delete(existing)
                sess.flush()

            correction = EmailCorrection(
                submission_id=submission_id,
                raw_text=result.get("raw_text"),
                row_count=len(result.get("corrections", [])),
                used_ai=bool(result.get("used_ai")),
            )
            sess.add(correction)
            sess.flush()

            for item in result.get("corrections", []):
                sess.add(EmailCorrectionRow(
                    correction_id=correction.id,
                    page_number=item.get("page_number"),
                    change_he=item.get("change_he", ""),
                    category=item.get("category", "drawing_change"),
                ))
            sess.commit()
            sess.refresh(correction)

            rows_out = [
                _EmailCorrectionRowOut(**r.to_dict()) for r in correction.rows
            ]
            return _EmailCorrectionOut(
                id=correction.id,
                submission_id=submission_id,
                row_count=correction.row_count,
                used_ai=bool(correction.used_ai),
                uploaded_at=correction.uploaded_at.isoformat(),
                rows=rows_out,
                raw_text=correction.raw_text,
                error=result.get("error"),
                error_message=result.get("error_message"),
            )

    # ── GET /submissions/{id}/email-corrections ────────────────────────

    @_subs_router.get(
        "/{submission_id}/email-corrections",
        response_model=_EmailCorrectionOut,
    )
    def get_email_corrections(submission_id: int) -> _EmailCorrectionOut:
        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            correction = sess.query(EmailCorrection).filter_by(
                submission_id=submission_id
            ).first()
            if correction is None:
                raise HTTPException(
                    404,
                    "לא נמצאו תיקוני מייל עבור הגשה זו",
                )
            rows_out = [
                _EmailCorrectionRowOut(**r.to_dict()) for r in correction.rows
            ]
            return _EmailCorrectionOut(
                id=correction.id,
                submission_id=submission_id,
                row_count=correction.row_count,
                used_ai=bool(correction.used_ai),
                uploaded_at=correction.uploaded_at.isoformat(),
                rows=rows_out,
            )

    # ── POST /submissions/{id}/upload-meeting-pdf ──────────────────────

    _MAX_MEETING_PDF_BYTES = 20 * 1024 * 1024  # 20 MB

    class _MeetingRowOut(BaseModel):
        id: int
        extraction_id: int
        row_type: str
        topic_he: str
        decision_he: str
        responsible_he: str | None = None
        deadline_he: str | None = None

    class _MeetingExtractionOut(BaseModel):
        id: int
        submission_id: int
        row_count: int
        used_ai: bool
        uploaded_at: str
        rows: list[_MeetingRowOut]
        raw_text: str | None = None
        error: str | None = None
        error_message: str | None = None
        truncation_warning: str | None = None

    @_subs_router.post(
        "/{submission_id}/upload-meeting-pdf",
        response_model=_MeetingExtractionOut,
        status_code=201,
    )
    async def upload_meeting_pdf(
        submission_id: int,
        pdf: UploadFile = File(...),
    ) -> _MeetingExtractionOut:
        """Extract action items and decisions from a meeting-notes PDF."""
        from .meeting_extract import extract_meeting_notes  # noqa: PLC0415
        from .models import MeetingExtraction, MeetingRow  # noqa: PLC0415

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")

        content = await pdf.read(_MAX_MEETING_PDF_BYTES + 1)
        if len(content) > _MAX_MEETING_PDF_BYTES:
            raise HTTPException(
                413,
                f"קובץ ה-PDF גדול מדי — מקסימום {_MAX_MEETING_PDF_BYTES // (1024*1024)} MB",
            )

        result = extract_meeting_notes(content)

        with _session() as sess:
            existing = sess.query(MeetingExtraction).filter_by(
                submission_id=submission_id
            ).first()
            if existing:
                sess.delete(existing)
                sess.flush()

            extraction = MeetingExtraction(
                submission_id=submission_id,
                raw_text=result.get("raw_text"),
                row_count=len(result.get("items", [])),
                used_ai=bool(result.get("used_ai")),
            )
            sess.add(extraction)
            sess.flush()

            for item in result.get("items", []):
                row = MeetingRow(
                    extraction_id=extraction.id,
                    row_type=item.get("row_type", "action_item"),
                    topic_he=item.get("topic_he", ""),
                    decision_he=item.get("decision_he", ""),
                    responsible_he=item.get("responsible_he"),
                    deadline_he=item.get("deadline_he"),
                )
                sess.add(row)
            sess.flush()

            rows_out = [_MeetingRowOut(**r.to_dict()) for r in extraction.rows]
            return _MeetingExtractionOut(
                id=extraction.id,
                submission_id=submission_id,
                row_count=extraction.row_count,
                used_ai=bool(extraction.used_ai),
                uploaded_at=extraction.uploaded_at.isoformat(),
                rows=rows_out,
                raw_text=result.get("raw_text"),
                error=result.get("error"),
                error_message=result.get("error_message"),
                truncation_warning=result.get("truncation_warning"),
            )

    # ── GET /submissions/{id}/meeting-notes ────────────────────────────

    @_subs_router.get(
        "/{submission_id}/meeting-notes",
        response_model=_MeetingExtractionOut,
    )
    def get_meeting_notes(submission_id: int) -> _MeetingExtractionOut:
        from .models import MeetingExtraction  # noqa: PLC0415

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            extraction = sess.query(MeetingExtraction).filter_by(
                submission_id=submission_id
            ).first()
            if extraction is None:
                raise HTTPException(404, "לא נמצא סיכום ישיבה עבור הגשה זו")
            rows_out = [_MeetingRowOut(**r.to_dict()) for r in extraction.rows]
            return _MeetingExtractionOut(
                id=extraction.id,
                submission_id=submission_id,
                row_count=extraction.row_count,
                used_ai=bool(extraction.used_ai),
                uploaded_at=extraction.uploaded_at.isoformat(),
                rows=rows_out,
            )

    # ── POST /submissions/{id}/verify-claims ───────────────────────────

    class _VerifyClaimIn(BaseModel):
        page_number: int
        claim_text: str
        verification_question: str

    class _VerifyClaimOut(BaseModel):
        page_number: int
        claim_text: str
        verification_question: str
        verified: str  # "yes" / "no" / "unclear"
        evidence: str
        status: str    # "תקין" / "דורש תיקון" / "דורש בירור"

    class _VerifyClaimsBody(BaseModel):
        claims: list[_VerifyClaimIn]

    @_subs_router.post(
        "/{submission_id}/verify-claims",
        response_model=list[_VerifyClaimOut],
    )
    async def verify_claims(
        submission_id: int,
        body: _VerifyClaimsBody,
    ) -> list[_VerifyClaimOut]:
        """Verify architect claimed changes via Gemini vision on specific pages."""
        from pathlib import Path as _Path  # noqa: PLC0415
        from compliance_engine.vision_verify import verify_claimed_changes  # noqa: PLC0415

        with _session() as sess:
            sub = sess.get(Submission, submission_id)
            if sub is None:
                raise HTTPException(404, f"submission {submission_id} not found")
            pdf_path = sub.pdf_path

        if not pdf_path:
            raise HTTPException(400, "להגשה זו אין קובץ PDF")

        claims_dicts = [c.model_dump() for c in body.claims]
        results = verify_claimed_changes(_Path(pdf_path), claims_dicts)
        return [_VerifyClaimOut(**r) for r in results]

    return _projects_subs_router, _subs_router


# ─────────────────────────────────────────────────────────────────────────────
# Range parsing + chunked file streaming (module-level helpers — testable)
# ─────────────────────────────────────────────────────────────────────────────

# RFC 7233-style Range header. We accept ONLY single byte ranges:
#   bytes=START-END   → both ends present
#   bytes=START-      → from START to EOF
#   bytes=-SUFFIX     → last SUFFIX bytes
_RANGE_RE = re.compile(r"^\s*bytes\s*=\s*(\d*)\s*-\s*(\d*)\s*$")


def _parse_range(header_value: str, file_size: int) -> tuple[int, int] | None:
    """Return (start, end) inclusive byte offsets, or None if unsatisfiable."""
    m = _RANGE_RE.match(header_value)
    if not m:
        return None
    raw_start, raw_end = m.group(1), m.group(2)

    if raw_start == "" and raw_end == "":
        return None  # "bytes=-" malformed
    if raw_start == "":
        # Suffix range: last N bytes.
        suffix = int(raw_end)
        if suffix == 0:
            return None
        start = max(0, file_size - suffix)
        end = file_size - 1
    else:
        start = int(raw_start)
        if start >= file_size:
            return None  # 416 path
        end = int(raw_end) if raw_end != "" else file_size - 1
        end = min(end, file_size - 1)
    if start > end:
        return None
    return start, end


def _stream_file_range(path: Path, start: int, end: int, chunk_size: int) -> Iterator[bytes]:
    """Yield bytes [start..end] inclusive from `path` in chunks. Never loads
    more than `chunk_size` bytes at a time into memory."""
    remaining = end - start + 1
    with path.open("rb") as f:
        f.seek(start)
        while remaining > 0:
            chunk = f.read(min(chunk_size, remaining))
            if not chunk:
                return
            remaining -= len(chunk)
            yield chunk


def _parse_response_xlsx(path: Path) -> list[dict]:
    """Parse the architect-filled response Excel produced by excel_export.py.

    Layout (from excel_export.COLUMNS):
      Row 1 — read-only warning banner (merged, skipped)
      Row 2 — column headers (validated before any data is read)
      Row 3+ — data rows

    Expected 0-based column indices:
      5  נושא
      6  סטטוס ממצא
      7  תיאור / פעולה נדרשת
      8  סטטוס טיפול        (architect-editable)
      9  הערות האדריכל       (architect-editable)
      10 source_id           (hidden round-trip key — critical)

    Raises ValueError with a Hebrew message when:
      - the header row is missing entirely
      - source_id column cannot be found
      - any other critical column (סטטוס טיפול, הערות האדריכל) is missing

    If columns are found but at different positions than expected (i.e. an
    architect shifted a column), the found positions are used and a warning
    is logged — no data is silently mis-mapped.
    """
    import openpyxl

    # Expected 0-based column indices, keyed by the exact header text.
    _EXPECTED: dict[str, int] = {
        "נושא": 5,
        "סטטוס ממצא": 6,
        "תיאור / פעולה נדרשת": 7,
        "סטטוס טיפול": 8,
        "הערות האדריכל": 9,
        "source_id": 10,
    }
    _CRITICAL = {"source_id", "סטטוס טיפול", "הערות האדריכל"}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # ── Step 1: read and validate the header row (row 2) ──────────────
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if header_row is None:
        wb.close()
        raise ValueError("מבנה הקובץ אינו תואם — לא נמצאה שורת כותרת")

    # Map header text → 0-based column index; try exact match first,
    # then strip-normalised match (handles stray spaces or BOM).
    actual: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            actual[str(cell).strip()] = idx

    col_map: dict[str, int] = {}
    shifted: list[str] = []
    missing: list[str] = []

    for name, expected_idx in _EXPECTED.items():
        if name in actual:
            found_idx = actual[name]
            col_map[name] = found_idx
            if found_idx != expected_idx:
                shifted.append(name)
        else:
            # Fuzzy fallback: case-insensitive + whitespace-normalised search.
            name_norm = " ".join(name.lower().split())
            match = next(
                (k for k in actual if " ".join(k.lower().split()) == name_norm), None
            )
            if match is not None:
                found_idx = actual[match]
                col_map[name] = found_idx
                if found_idx != expected_idx:
                    shifted.append(name)
            else:
                missing.append(name)

    if "source_id" in missing:
        wb.close()
        raise ValueError(
            "עמודת מזהה (source_id) חסרה. "
            "נא להשתמש בקובץ האקסל המקורי שהופק מהמערכת."
        )

    critical_missing = [m for m in missing if m in _CRITICAL]
    if critical_missing:
        wb.close()
        raise ValueError(
            "מבנה הקובץ אינו תואם. ודאי שלא שונו עמודות באקסל. "
            f"העמודות הבאות חסרות: {', '.join(critical_missing)}"
        )

    if shifted:
        log.warning(
            "response xlsx: %d column(s) at unexpected positions: %s — "
            "using found positions",
            len(shifted), shifted,
        )

    # Resolve final indices (fall back to expected if a non-critical column was missing).
    i_sid = col_map["source_id"]
    i_topic = col_map.get("נושא", _EXPECTED["נושא"])
    i_status = col_map.get("סטטוס ממצא", _EXPECTED["סטטוס ממצא"])
    i_desc = col_map.get("תיאור / פעולה נדרשת", _EXPECTED["תיאור / פעולה נדרשת"])
    i_treatment = col_map["סטטוס טיפול"]
    i_notes = col_map["הערות האדריכל"]

    min_cols = max(i_sid, i_topic, i_status, i_desc, i_treatment, i_notes) + 1

    # ── Step 2: parse data rows (row 3 onward) ────────────────────────
    def _cell(row: tuple, idx: int) -> str | None:
        if idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip() or None
        return None

    result: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < min_cols:
            continue
        raw_sid = row[i_sid]
        source_id = str(raw_sid).strip() if raw_sid is not None else ""
        if not source_id or source_id.lower() == "none":
            continue
        result.append({
            "source_id": source_id,
            "topic_he": _cell(row, i_topic),
            "finding_status": _cell(row, i_status),
            "description": _cell(row, i_desc),
            "treatment_status": _cell(row, i_treatment),
            "architect_notes": _cell(row, i_notes),
        })
    wb.close()
    return result
